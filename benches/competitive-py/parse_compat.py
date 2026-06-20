"""PARSE conformance + perf: turbo-xlsx (PyO3) vs openpyxl.

openpyxl writes a DEFLATE-compressed ``.xlsx`` (the Excel-style shape turbo's
own STORED writer never produces), then both turbo and openpyxl read it back. We
diff every cell after a shared canonicalization to prove turbo is not "outputting
garbage", then time a large read both ways.

Setup (from this directory):
    python3 -m venv .venv && . .venv/bin/activate
    pip install -r requirements.txt maturin
    maturin develop --manifest-path ../../crates/turbo-xlsx-py/Cargo.toml --features parse
    python parse_compat.py
"""

from __future__ import annotations

import io
import json
import time

import openpyxl

import turbo_xlsx as tx

if not hasattr(tx, "parse"):
    raise SystemExit(
        "turbo_xlsx.parse is unavailable — rebuild with the parse feature:\n"
        "  maturin develop --manifest-path ../../crates/turbo-xlsx-py/Cargo.toml --features parse"
    )


def sample_grid() -> list[list[object]]:
    # Unicode, embedded commas/quotes, empty strings, zero, negatives, large and
    # fractional numbers, booleans. Dates are excluded on purpose (timezone noise
    # — serial->ISO conversion is proven exactly in the core unit tests).
    return [
        ["id", "name", "amount", "ratio", "active", "score", "note"],
        [1, "Alice", 1234.56, 0.125, True, 88, "repeat"],
        [2, 'O’Brien, "Bob"', -42, 0, False, -1.5, 'commas, "quotes" ☃'],
        [3, "", 0.1, 1, True, 0, "repeat"],
        [4, "Zoë", 9999999.99, 0.9999, False, 100000, "tail"],
    ]


def via_openpyxl(grid: list[list[object]]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    for row in grid:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)  # openpyxl always DEFLATEs
    return buf.getvalue()


def read_openpyxl(data: bytes) -> list[list[object]]:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    ws = wb[wb.sheetnames[0]]
    return [list(r) for r in ws.iter_rows(values_only=True)]


def turbo_grid(data: bytes) -> list[list[object]]:
    return json.loads(tx.parse(data, format="json"))["sheets"][0]["rows"]


def canon(v: object) -> str:
    if v is None or v == "":
        return ""
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        if abs(v) < 1e-9:
            return "0"
        f = float(v)
        return str(int(f)) if f == int(f) else str(round(f, 6))
    return str(v)


def diff(a: list[list[object]], b: list[list[object]]) -> list[tuple[int, int, str, str]]:
    misses: list[tuple[int, int, str, str]] = []
    for r in range(max(len(a), len(b))):
        ra = a[r] if r < len(a) else []
        rb = b[r] if r < len(b) else []
        for c in range(max(len(ra), len(rb))):
            ca = canon(ra[c]) if c < len(ra) else ""
            cb = canon(rb[c]) if c < len(rb) else ""
            if ca != cb:
                misses.append((r, c, ca, cb))
    return misses


def perf_grid(rows: int) -> list[list[object]]:
    grid: list[list[object]] = [["id", "label", "a", "b", "c", "d", "e", "flag"]]
    for i in range(rows):
        grid.append([i, f"row-{i}", i * 1.5, i - 7, i % 100, i / 3, -i, i % 2 == 0])
    return grid


def median(xs: list[float]) -> float:
    return sorted(xs)[len(xs) // 2]


def time_read(fn, data: bytes, iters: int) -> float:
    times = []
    for _ in range(iters):
        t0 = time.perf_counter()
        fn(data)
        times.append((time.perf_counter() - t0) * 1000)
    return median(times)


def main() -> None:
    grid = sample_grid()
    data = via_openpyxl(grid)
    misses = diff(turbo_grid(data), read_openpyxl(data))
    cells = len(grid) * len(grid[0])
    if misses:
        print(f"✗ openpyxl  {len(misses)}/{cells} cell(s) differ:")
        for r, c, a, b in misses[:8]:
            print(f"    [r{r},c{c}] turbo={a!r} ref={b!r}")
        raise SystemExit(1)
    print(f"✓ openpyxl  {cells} cells parsed identically (DEFLATE, {len(data)}B)")

    csv = tx.parse(data, format="csv")
    md = tx.parse(data, format="md")
    print(f"serializers: csv {len(csv)}B, markdown {len(md)}B")

    print("\nparse perf (read DEFLATEd file -> value grid):")
    for rows in (1_000, 50_000):
        buf = via_openpyxl(perf_grid(rows))
        iters = 5 if rows >= 50_000 else 20
        turbo_grid(buf)  # warm
        read_openpyxl(buf)
        t = time_read(turbo_grid, buf, iters)
        o = time_read(read_openpyxl, buf, iters)
        print(
            f"  {rows:>6} rows ({len(buf) // 1024}KB):  "
            f"turbo {t:6.2f}ms   openpyxl {o:7.2f}ms   -> {o / t:.1f}x faster"
        )

    print("\nturbo-xlsx parses openpyxl's DEFLATEd output cell-for-cell. ✓")


if __name__ == "__main__":
    main()
