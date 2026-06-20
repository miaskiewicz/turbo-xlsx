"""Competitive performance: turbo-xlsx (PyO3) vs openpyxl vs XlsxWriter.

Same logical workloads as the Node harness — a styled 1k-row sheet and a 50k-row
sheet — written by each library, timed, and sized. turbo-xlsx uses its streaming
``write_rows_json`` throughput path (a chunk stringified in C, parsed once in
Rust), the same trick the Node binding uses to skip the per-cell FFI walk.

Setup (from this directory):
    python3 -m venv .venv && . .venv/bin/activate
    pip install -r requirements.txt maturin
    maturin develop --manifest-path ../../crates/turbo-xlsx-py/Cargo.toml
    python perf.py
"""

from __future__ import annotations

import io
import time

import openpyxl
import xlsxwriter
from openpyxl.cell import WriteOnlyCell

import turbo_xlsx as tx

CURRENCY_NUMFMT = '"$"#,##0.00;[Red]("$"#,##0.00)'
HEADER_FILL = "DDDDDD"


def workload(rows: int, cols: int) -> list[list[tuple[str, object]]]:
    """A neutral grid: a bold header band then label + currency (minor units)."""
    grid: list[list[tuple[str, object]]] = []
    grid.append([("header", f"Col {c}") for c in range(cols)])
    for r in range(rows):
        row: list[tuple[str, object]] = [("string", f"Empleado {r}")]
        row.extend(("currency", (r * c + 1) * 100) for c in range(1, cols))
        grid.append(row)
    return grid


# ---- turbo-xlsx (streaming write_rows_json) --------------------------------
def _turbo_cell(kind: str, value: object) -> dict:
    if kind == "currency":
        return {"type": "currency", "value": value, "currency": {"code": "MXN", "locale": "es-MX"}}
    style = {"font": {"bold": True}, "fill": "#dddddd"} if kind == "header" else None
    cell = {"type": "string", "value": value}
    if style:
        cell["style"] = style
    return cell


def turbo_write(grid: list[list[tuple[str, object]]]) -> bytes:
    # Typed-table fast path: the header row (styled strings) goes through
    # write_row; the uniform data rows go through write_table as bare scalar
    # values + a per-column type spec — no per-cell dicts, no JSON.
    w = tx.create_writer({"locale": "es-MX"})
    w.start_sheet({"name": "Bench"})
    header = grid[0]
    w.write_row({"cells": [_turbo_cell("header", v) for (_, v) in header]})
    columns = [{"type": "string"}] + [
        {"type": "currency", "currency": {"code": "MXN", "locale": "es-MX"}}
        for _ in range(len(header) - 1)
    ]
    rows = [[v for (_, v) in row] for row in grid[1:]]
    w.write_table(columns, rows)
    w.end_sheet()
    data, _ = w.finish()
    return data


# ---- openpyxl (write_only mode, the fast path) -----------------------------
def openpyxl_write(grid: list[list[tuple[str, object]]]) -> bytes:
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    for row in grid:
        cells = []
        for kind, value in row:
            if kind == "currency":
                c = WriteOnlyCell(ws, value=value / 100)
                c.number_format = CURRENCY_NUMFMT
            elif kind == "header":
                c = WriteOnlyCell(ws, value=value)
                c.font = openpyxl.styles.Font(bold=True)
            else:
                c = WriteOnlyCell(ws, value=value)
            cells.append(c)
        ws.append(cells)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---- XlsxWriter -------------------------------------------------------------
def xlsxwriter_write(grid: list[list[tuple[str, object]]]) -> bytes:
    buf = io.BytesIO()
    wb = xlsxwriter.Workbook(buf, {"in_memory": True})
    ws = wb.add_worksheet("Bench")
    cur_fmt = wb.add_format({"num_format": CURRENCY_NUMFMT})
    hdr_fmt = wb.add_format({"bold": True, "bg_color": f"#{HEADER_FILL}"})
    for r, row in enumerate(grid):
        for c, (kind, value) in enumerate(row):
            if kind == "currency":
                ws.write_number(r, c, value / 100, cur_fmt)
            elif kind == "header":
                ws.write_string(r, c, value, hdr_fmt)
            else:
                ws.write_string(r, c, value)
    wb.close()
    return buf.getvalue()


ADAPTERS = [
    ("turbo-xlsx", turbo_write),
    ("openpyxl", openpyxl_write),
    ("XlsxWriter", xlsxwriter_write),
]
WORKLOADS = [("1k x 20 styled", 1_000, 20, 5), ("50k x 30", 50_000, 30, 3)]


def measure(fn, grid, reps: int) -> tuple[float, int]:
    times = []
    size = 0
    for _ in range(reps):
        t0 = time.perf_counter()
        out = fn(grid)
        times.append(time.perf_counter() - t0)
        size = len(out)
    times.sort()
    return times[len(times) // 2], size


def fmt_ms(s: float) -> str:
    return f"{s * 1000:.1f} ms" if s < 1 else f"{s:.2f} s"


def fmt_bytes(b: int) -> str:
    return f"{b / 1e6:.1f} MB" if b >= 1e6 else f"{b / 1e3:.0f} KB"


def main() -> None:
    lines = ["# Python competitive performance\n"]
    fastest_all = True
    for name, rows, cols, reps in WORKLOADS:
        grid = workload(rows, cols)
        print(f"\n## {name}  (median of {reps})")
        print("library".ljust(14) + "time".rjust(12) + "output".rjust(12))
        lines.append(f"\n## {name}  (median of {reps})\n")
        lines.append("| library | time | output |\n|---|---|---|")
        results = []
        for lib, fn in ADAPTERS:
            ms, size = measure(fn, grid, reps)
            results.append((lib, ms, size))
            print(lib.ljust(14) + fmt_ms(ms).rjust(12) + fmt_bytes(size).rjust(12))
            lines.append(f"| {lib} | {fmt_ms(ms)} | {fmt_bytes(size)} |")
        winner = min(results, key=lambda r: r[1])[0]
        if winner != "turbo-xlsx":
            fastest_all = False
        print(f"  -> fastest: {winner}")
    with open("RESULTS.py.md", "w") as f:
        f.write("\n".join(lines) + "\n")
    verdict = "turbo-xlsx is FASTEST on every workload" if fastest_all else "turbo-xlsx is NOT fastest everywhere"
    print(f"\n{verdict}")
    raise SystemExit(0 if fastest_all else 1)


if __name__ == "__main__":
    main()
